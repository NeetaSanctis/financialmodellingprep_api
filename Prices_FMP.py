#!/usr/bin/env python
try:
    # For Python 3.0 and later
    from urllib.request import urlopen
except ImportError:
    # Fall back to Python 2's urllib2
    from urllib2 import urlopen

import certifi
import json
import os
import openpyxl
import pyodbc
import yfinance as yf
import pandas as pd
import csv
from pandas_datareader import data as pdr
yf.pdr_override() # <== that's all it takes :-)
import warnings
warnings.filterwarnings("ignore")
p = os.path.join (os.path.expanduser ("~"), "desktop")
print(p)
path = p + "\MORNING.xlsx" 
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
for i in range(2, m_row+1):
    cell_obj = sheet_obj.cell(row = i, column = 2)    
    ticker = cell_obj.value
    print(i)
    try:
        def get_jsonparsed_data(url):
            response = urlopen(url, cafile=certifi.where())
            data = response.read().decode("utf-8")      
            f= data.split(',')
            
            p = f[1].split('{')        
            date1= p[1].split(':')
            c4 = sheet_obj.cell(row = i, column = 3)
            invalid_chars = '<>:"/\\|?*'
            for char in invalid_chars:
                date1[1] = date1[1].replace(char, '')            
            c4.value = date1[1].strip()
            
            close1= f[5].split(':')   
            c4 = sheet_obj.cell(row = i, column = 6)
            c4.value = close1[1]

            o = f[15].split('{')
            date2= o[1].split(':')
            c4 = sheet_obj.cell(row = i, column = 7)
            invalid_chars = '<>:"/\\|?*'
            for char in invalid_chars:
                date2[1] = date2[1].replace(char, '')            
            c4.value = date2[1].strip()            

            close2= f[19].split(':')   
            c4 = sheet_obj.cell(row = i, column = 10)
            c4.value = close2[1]
           
            wb_obj.save(path)
            return json.loads(data)
    except Exception:
        wb_obj.save(path)
        pass

    today = datetime.date.today()
    url = ("https://financialmodelingprep.com/api/v3/historical-price-full/" + ticker + "?limit=10&from=2025-02-18&to="+str(today)+"&apikey=aa91b6fd2bd83cac2f3364e9f4d07e86")
    get_jsonparsed_data(url)

print(x.strftime("Prices Fetched"))
